# APP.PY - MINI CRM THU CƯỚC KHDN VNPT
# Company-based CRM v2: DS giao -> lọc nhân viên -> TN08 -> gom công ty -> gọi theo công ty

import io, re, sqlite3, unicodedata
from datetime import datetime
from decimal import Decimal, InvalidOperation

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

APP_CONFIG = {
    "app_title": "📞 Mini-CRM Thu Cước KHDN - Company-based v2",
    "db_path": "crm_vnpt_company_crm_v2.db",
    "default_staff": "Vương Thanh Thuận",
    "defaults": {"collection_team": "KD - DN", "route_name": "", "sl_hd": 1},
    "message": {
        "unit_name": "VNPT Long Thành- Nhơn Trạch",
        "website": "https://vnptdongnai.vn/",
        "payment_deadline": "ngày 15",
        "staff_name": "Thuận",
        "staff_phone": "0837892579",
        "default_period": "05",
        "template": """{unit_name} thông báo:
Đã có thông báo cước kỳ cước tháng {period}. Anh/chị truy cập trang {website} để lấy thông báo cước và hóa đơn của công ty.
Anh/chị vui lòng thanh toán cước trước {payment_deadline}. Sau ngày 15 hệ thống sẽ tự động đưa lưới khóa khi ghi nhận còn tồn nợ.
Liên hệ nhân viên kinh doanh: {staff_name}: {staff_phone} để được hỗ trợ gạch nợ sớm nhất sau khi thanh toán.
Nếu đã thanh toán cước, anh/chị vui lòng bỏ qua tin nhắn trên.
Trân trọng."""
    },
    "assignment_columns": {
        "collection_team": ["tổ", "to", "tổ thu cước", "to thu cuoc"],
        "route_name": ["tuyến kỹ thuật", "tuyen ky thuat", "tuyến", "tuyen"],
        "staff_name": ["nhân viên thu cước", "nhan vien thu cuoc", "nhân viên", "nhan vien", "nv thu", "người phụ trách", "nguoi phu trach"],
        "sl_hd": ["sl hđ", "sl hd", "sl", "số lượng", "so luong"],
        "ma_tt": ["mã thanh toán", "ma thanh toan", "mã tt", "ma_tt", "ma tt"],
        "customer_name": ["tên thanh toán", "ten thanh toan", "tên khách hàng", "ten khach hang", "khách hàng", "khach hang"],
        "generated_amount": ["tiền phát sinh", "tien phat sinh", "phát sinh", "phat sinh"],
        "phone": ["số dt liên hệ", "so dt lien he", "số điện thoại", "so dien thoai", "điện thoại", "dien thoai", "sdt", "phone"],
        "serial": ["số seri", "so seri", "seri", "serial"],
        "address": ["địa chỉ thanh toán", "dia chi thanh toan", "địa chỉ", "dia chi"],
        "old_status_note": ["ghi chú tình trạng", "ghi chu tinh trang", "tình trạng", "tinh trang"],
    },
    "tn08_columns": {
        "ma_tt": ["ma_tt", "mã tt", "ma tt", "mã thanh toán", "ma thanh toan"],
        "debt_amount": ["total nợ thu vét", "total no thu vet", "total nợ thu vét", "nợ thu vét", "no thu vet", "cn tn08", "tiền nợ", "tien no"],
        "customer_name_tn08": ["tên khách hàng", "ten khach hang", "tên thanh toán", "ten thanh toan"],
        "address_tn08": ["địa chỉ kh", "dia chi kh", "địa chỉ", "dia chi"],
    },
    "paid_columns": {
        "ma_tt": ["ma_tt", "mã tt", "ma tt", "mã thanh toán", "ma thanh toan"],
        "paid_amount": ["số tiền", "so tien", "tiền đóng", "tien dong", "số tiền đã đóng"],
        "paid_date": ["ngày đóng", "ngay dong", "ngày thanh toán", "ngay thanh toan"],
    },
    "contact_results": ["Đã gọi - hẹn thanh toán", "Đã gọi - xin số kế toán mới", "Đã gọi - khách báo đã đóng", "Không nghe máy", "Sai số điện thoại", "Không có số điện thoại", "Cần kiểm tra lại"],
    "status": {
        "paid": ("🟢", "Đã đóng"),
        "partial_paid": ("🔵", "Đã đóng một phần"),
        "contacted": ("🟡", "Đã liên hệ chưa đóng"),
        "uncontacted": ("🔴", "Chưa liên hệ"),
        "no_phone": ("🔴", "Thiếu số điện thoại"),
        "need_check": ("⚪", "Cần kiểm tra dữ liệu"),
    }
}

def normalize_text(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).strip()
    if text.lower() in ["nan", "nat", "none", "null"]:
        return ""
    return re.sub(r"\s+", " ", text)

def remove_accents(value) -> str:
    text = normalize_text(value).lower()
    text = unicodedata.normalize("NFD", text)
    return "".join(ch for ch in text if unicodedata.category(ch) != "Mn")

def normalize_key(value) -> str:
    return re.sub(r"[^a-z0-9]+", "_", remove_accents(value)).strip("_")

def normalize_ma_tt(value) -> str:
    text = normalize_text(value)
    if text.endswith(".0"):
        text = text[:-2]
    return text.upper()

def normalize_phone(value) -> str:
    text = normalize_text(value)
    if text.endswith(".0"):
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    if not digits:
        return ""
    if len(digits) == 9:
        digits = "0" + digits
    return digits

def parse_money(value) -> float:
    if value is None or value == "":
        return 0.0
    try:
        if pd.isna(value):
            return 0.0
    except Exception:
        pass
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_text(value)
    if text in ["-", "–", "—"]:
        return 0.0
    text = text.replace("đ", "").replace("Đ", "").replace("VND", "").replace("vnd", "").replace(" ", "")
    if "." in text and "," in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return float(Decimal(text))
    except (InvalidOperation, ValueError):
        return 0.0

def format_money(value) -> str:
    try:
        return f"{float(value):,.0f} đồng".replace(",", ".")
    except Exception:
        return "0 đồng"

def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def staff_key(value) -> str:
    return normalize_key(value)

def company_key(value) -> str:
    return normalize_key(value)

def find_col(df: pd.DataFrame, candidates: list[str]):
    normalized = {normalize_key(col): col for col in df.columns}
    for c in candidates:
        k = normalize_key(c)
        if k in normalized:
            return normalized[k]
    return None

def extract_email(text):
    text = normalize_text(text)
    match = re.search(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", text)
    return match.group(0) if match else ""

def get_conn():
    conn = sqlite3.connect(APP_CONFIG["db_path"], check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def init_db(conn):
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS import_batches (
            batch_id INTEGER PRIMARY KEY AUTOINCREMENT,
            import_type TEXT NOT NULL,
            file_name TEXT,
            staff_filter TEXT,
            row_count INTEGER DEFAULT 0,
            total_amount REAL DEFAULT 0,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS work_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id INTEGER NOT NULL,
            company_key TEXT NOT NULL,
            company_name TEXT NOT NULL,
            staff_name TEXT,
            collection_team TEXT,
            route_name TEXT,
            sl_hd INTEGER DEFAULT 1,
            ma_tt TEXT NOT NULL,
            customer_name TEXT,
            phone TEXT,
            address TEXT,
            generated_amount REAL DEFAULT 0,
            debt_amount REAL DEFAULT 0,
            tn08_customer_name TEXT,
            tn08_address TEXT,
            serial TEXT,
            old_status_note TEXT,
            data_status TEXT,
            created_at TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_work_items_batch_company ON work_items(batch_id, company_key);
        CREATE INDEX IF NOT EXISTS idx_work_items_batch_ma_tt ON work_items(batch_id, ma_tt);
        CREATE TABLE IF NOT EXISTS paid_updates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id INTEGER NOT NULL,
            ma_tt TEXT NOT NULL,
            paid_amount REAL DEFAULT 0,
            paid_date TEXT,
            created_at TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_paid_updates_ma_tt ON paid_updates(ma_tt);
        CREATE TABLE IF NOT EXISTS interactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_key TEXT NOT NULL,
            result TEXT NOT NULL,
            promised_payment_date TEXT,
            note TEXT,
            created_by TEXT,
            created_at TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_interactions_company ON interactions(company_key);
        CREATE TABLE IF NOT EXISTS contacts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_key TEXT NOT NULL,
            contact_value TEXT NOT NULL,
            contact_person TEXT,
            role TEXT,
            note TEXT,
            created_at TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_contacts_company ON contacts(company_key);
        CREATE TABLE IF NOT EXISTS sent_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_key TEXT NOT NULL,
            message_text TEXT NOT NULL,
            created_by TEXT,
            created_at TEXT NOT NULL
        );
    """)
    conn.commit()

def create_batch(conn, import_type, file_name, staff_filter, row_count, total_amount):
    cur = conn.execute("""
        INSERT INTO import_batches
        (import_type, file_name, staff_filter, row_count, total_amount, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (import_type, file_name, staff_filter, row_count, total_amount, now_str()))
    conn.commit()
    return int(cur.lastrowid)

def get_latest_batch_id(conn, import_type):
    row = conn.execute("""
        SELECT batch_id FROM import_batches
        WHERE import_type = ?
        ORDER BY batch_id DESC
        LIMIT 1
    """, (import_type,)).fetchone()
    return int(row["batch_id"]) if row else None

def insert_work_items(conn, batch_id, df):
    rows = []
    for _, r in df.iterrows():
        rows.append((
            batch_id, r.get("company_key", ""), r.get("company_name", ""),
            r.get("staff_name", ""), r.get("collection_team", ""), r.get("route_name", ""),
            int(r.get("sl_hd", 1) or 1), r.get("ma_tt", ""), r.get("customer_name", ""),
            r.get("phone", ""), r.get("address", ""), float(r.get("generated_amount", 0) or 0),
            float(r.get("debt_amount", 0) or 0), r.get("customer_name_tn08", ""),
            r.get("address_tn08", ""), r.get("serial", ""), r.get("old_status_note", ""),
            r.get("data_status", ""), now_str()
        ))
    conn.executemany("""
        INSERT INTO work_items
        (batch_id, company_key, company_name, staff_name, collection_team, route_name, sl_hd,
         ma_tt, customer_name, phone, address, generated_amount, debt_amount,
         tn08_customer_name, tn08_address, serial, old_status_note, data_status, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, rows)
    conn.commit()

def insert_paid_updates(conn, batch_id, df):
    rows = []
    for _, r in df.iterrows():
        rows.append((batch_id, r.get("ma_tt", ""), float(r.get("paid_amount", 0) or 0), r.get("paid_date", ""), now_str()))
    conn.executemany("""
        INSERT INTO paid_updates
        (batch_id, ma_tt, paid_amount, paid_date, created_at)
        VALUES (?, ?, ?, ?, ?)
    """, rows)
    conn.commit()

def add_interaction(conn, company_key_value, result, promised_payment_date, note, created_by):
    conn.execute("""
        INSERT INTO interactions
        (company_key, result, promised_payment_date, note, created_by, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (company_key_value, result, promised_payment_date, note, created_by, now_str()))
    conn.commit()

def add_contact(conn, company_key_value, contact_value, contact_person, role, note):
    conn.execute("""
        INSERT INTO contacts
        (company_key, contact_value, contact_person, role, note, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (company_key_value, contact_value, contact_person, role, note, now_str()))
    conn.commit()

def add_sent_message(conn, company_key_value, message_text, created_by):
    conn.execute("""
        INSERT INTO sent_messages
        (company_key, message_text, created_by, created_at)
        VALUES (?, ?, ?, ?)
    """, (company_key_value, message_text, created_by, now_str()))
    conn.commit()

def list_sheets(uploaded_file):
    xls = pd.ExcelFile(uploaded_file)
    sheets = xls.sheet_names
    uploaded_file.seek(0)
    return sheets

def read_file(uploaded_file, sheet_name=None):
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        try:
            return pd.read_csv(uploaded_file)
        except UnicodeDecodeError:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, encoding="latin1")
    return pd.read_excel(uploaded_file, sheet_name=sheet_name)

def standardize_assignment(df_raw, fallback_staff=None):
    df = df_raw.copy()
    df.columns = [normalize_text(c) for c in df.columns]
    ma_col = find_col(df, APP_CONFIG["assignment_columns"]["ma_tt"])
    if not ma_col:
        raise ValueError("DS giao thiếu cột Mã thanh toán.")
    out = pd.DataFrame()
    out["ma_tt"] = df[ma_col]
    staff_col = find_col(df, APP_CONFIG["assignment_columns"]["staff_name"])
    out["staff_name"] = df[staff_col] if staff_col else (fallback_staff or APP_CONFIG["default_staff"])
    for field in ["collection_team", "route_name", "sl_hd", "customer_name", "generated_amount", "phone", "serial", "address", "old_status_note"]:
        col = find_col(df, APP_CONFIG["assignment_columns"][field])
        if col:
            out[field] = df[col]
        else:
            if field == "collection_team":
                out[field] = APP_CONFIG["defaults"]["collection_team"]
            elif field == "route_name":
                out[field] = APP_CONFIG["defaults"]["route_name"]
            elif field == "sl_hd":
                out[field] = APP_CONFIG["defaults"]["sl_hd"]
            else:
                out[field] = ""
    out["staff_name"] = out["staff_name"].apply(normalize_text)
    out["staff_key"] = out["staff_name"].apply(staff_key)
    out["ma_tt"] = out["ma_tt"].apply(normalize_ma_tt)
    out["collection_team"] = out["collection_team"].apply(normalize_text)
    out["route_name"] = out["route_name"].apply(normalize_text)
    out["sl_hd"] = out["sl_hd"].apply(lambda x: int(parse_money(x) or 1))
    out["customer_name"] = out["customer_name"].apply(normalize_text)
    out["phone"] = out["phone"].apply(normalize_phone)
    out["serial"] = out["serial"].apply(normalize_text)
    out["address"] = out["address"].apply(normalize_text)
    out["old_status_note"] = out["old_status_note"].apply(normalize_text)
    out["generated_amount"] = out["generated_amount"].apply(parse_money)
    out = out[out["ma_tt"].astype(str).str.len() > 0].copy()
    return out

def standardize_tn08(df_raw):
    df = df_raw.copy()
    df.columns = [normalize_text(c) for c in df.columns]
    ma_col = find_col(df, APP_CONFIG["tn08_columns"]["ma_tt"])
    debt_col = find_col(df, APP_CONFIG["tn08_columns"]["debt_amount"])
    if not ma_col:
        raise ValueError("TN08 thiếu cột MA_TT.")
    if not debt_col:
        raise ValueError("TN08 thiếu cột Total Nợ thu vét. App không dùng Tiền phát sinh làm tiền cần thu.")
    out = pd.DataFrame()
    out["ma_tt"] = df[ma_col]
    out["debt_amount"] = df[debt_col]
    for field in ["customer_name_tn08", "address_tn08"]:
        col = find_col(df, APP_CONFIG["tn08_columns"][field])
        out[field] = df[col] if col else ""
    out["ma_tt"] = out["ma_tt"].apply(normalize_ma_tt)
    out["debt_amount"] = out["debt_amount"].apply(parse_money)
    out["customer_name_tn08"] = out["customer_name_tn08"].apply(normalize_text)
    out["address_tn08"] = out["address_tn08"].apply(normalize_text)
    out = out[out["ma_tt"].astype(str).str.len() > 0].copy()
    out = out.groupby("ma_tt", as_index=False).agg({
        "debt_amount": "sum",
        "customer_name_tn08": "first",
        "address_tn08": "first",
    })
    return out

def standardize_paid(df_raw):
    df = df_raw.copy()
    df.columns = [normalize_text(c) for c in df.columns]
    ma_col = find_col(df, APP_CONFIG["paid_columns"]["ma_tt"])
    if not ma_col:
        raise ValueError("File đã đóng thiếu cột MA_TT / Mã thanh toán.")
    out = pd.DataFrame()
    out["ma_tt"] = df[ma_col]
    amount_col = find_col(df, APP_CONFIG["paid_columns"]["paid_amount"])
    out["paid_amount"] = df[amount_col] if amount_col else 0
    date_col = find_col(df, APP_CONFIG["paid_columns"]["paid_date"])
    out["paid_date"] = df[date_col] if date_col else ""
    out["ma_tt"] = out["ma_tt"].apply(normalize_ma_tt)
    out["paid_amount"] = out["paid_amount"].apply(parse_money)
    out["paid_date"] = out["paid_date"].apply(normalize_text)
    out = out[out["ma_tt"].astype(str).str.len() > 0].copy()
    return out

def build_working_dataset(df_assignment, df_tn08, selected_staff):
    selected_key = staff_key(selected_staff)
    df_staff = df_assignment[df_assignment["staff_key"] == selected_key].copy()
    if df_staff.empty:
        return pd.DataFrame()
    route_ma = set(df_staff["ma_tt"].dropna().astype(str))
    df_tn08_small = df_tn08[df_tn08["ma_tt"].isin(route_ma)].copy()
    merged = df_staff.merge(
        df_tn08_small[["ma_tt", "debt_amount", "customer_name_tn08", "address_tn08"]],
        on="ma_tt",
        how="left"
    )
    merged["debt_amount"] = merged["debt_amount"].fillna(0)
    merged["customer_name_tn08"] = merged["customer_name_tn08"].fillna("")
    merged["address_tn08"] = merged["address_tn08"].fillna("")
    merged["company_name"] = merged["customer_name"].where(merged["customer_name"].astype(str).str.len() > 0, merged["customer_name_tn08"])
    merged["company_name"] = merged["company_name"].apply(normalize_text)
    merged["company_key"] = merged["company_name"].apply(company_key)
    merged["data_status"] = merged["debt_amount"].apply(lambda x: "Có nợ theo TN08" if x > 0 else "Không có trong TN08 / nợ = 0 / cần kiểm tra")
    merged["debt_amount_display"] = merged["debt_amount"].apply(format_money)
    merged["generated_amount_display"] = merged["generated_amount"].apply(format_money)
    return merged

def get_current_lines(conn):
    latest_work_batch = get_latest_batch_id(conn, "working_dataset")
    latest_paid_batch = get_latest_batch_id(conn, "paid")
    if latest_work_batch is None:
        return pd.DataFrame()
    if latest_paid_batch is None:
        paid_cte = "SELECT '' AS ma_tt, 0 AS paid_amount WHERE 1=0"
        params = [latest_work_batch]
    else:
        paid_cte = """
            SELECT ma_tt, SUM(paid_amount) AS paid_amount
            FROM paid_updates
            WHERE batch_id = ?
            GROUP BY ma_tt
        """
        params = [latest_paid_batch, latest_work_batch]
    query = f"""
        WITH paid AS ({paid_cte}),
        last_interaction AS (
            SELECT i.*
            FROM interactions i
            INNER JOIN (
                SELECT company_key, MAX(created_at) AS max_created_at
                FROM interactions
                GROUP BY company_key
            ) x ON i.company_key = x.company_key AND i.created_at = x.max_created_at
        ),
        latest_contact AS (
            SELECT c.*
            FROM contacts c
            INNER JOIN (
                SELECT company_key, MAX(created_at) AS max_created_at
                FROM contacts
                GROUP BY company_key
            ) x ON c.company_key = x.company_key AND c.created_at = x.max_created_at
        )
        SELECT
            w.*,
            COALESCE(p.paid_amount, 0) AS paid_amount,
            li.result AS last_result,
            li.promised_payment_date,
            li.note AS last_note,
            li.created_by AS last_created_by,
            li.created_at AS last_contacted_at,
            COALESCE(NULLIF(lc.contact_value, ''), w.phone) AS current_phone
        FROM work_items w
        LEFT JOIN paid p ON p.ma_tt = w.ma_tt
        LEFT JOIN last_interaction li ON li.company_key = w.company_key
        LEFT JOIN latest_contact lc ON lc.company_key = w.company_key
        WHERE w.batch_id = ?
        ORDER BY w.debt_amount DESC
    """
    df = pd.read_sql_query(query, conn, params=params)
    if df.empty:
        return df
    df["company_display"] = df["company_name"].where(df["company_name"].astype(str).str.len() > 0, df["customer_name"])
    df["debt_amount_display"] = df["debt_amount"].apply(format_money)
    df["generated_amount_display"] = df["generated_amount"].apply(format_money)
    df["paid_amount_display"] = df["paid_amount"].apply(format_money)
    return df

def build_company_view(lines_df: pd.DataFrame):
    if lines_df.empty:
        return pd.DataFrame()
    grouped = lines_df.groupby("company_key", as_index=False).agg(
        company_name=("company_display", "first"),
        phone=("current_phone", lambda s: next((normalize_text(x) for x in s if normalize_text(x)), "")),
        address=("address", "first"),
        ma_tt_count=("ma_tt", "nunique"),
        total_debt=("debt_amount", "sum"),
        paid_line_count=("paid_amount", lambda s: int((s.astype(float) > 0).sum())),
        line_count=("ma_tt", "count"),
        last_result=("last_result", "first"),
        promised_payment_date=("promised_payment_date", "first"),
        last_note=("last_note", "first"),
        last_contacted_at=("last_contacted_at", "first"),
        route_name=("route_name", "first"),
        collection_team=("collection_team", "first"),
        staff_name=("staff_name", "first"),
    )
    def calc_company_status(row):
        if row["line_count"] > 0 and row["paid_line_count"] >= row["line_count"]:
            return "paid"
        if row["paid_line_count"] > 0:
            return "partial_paid"
        if float(row["total_debt"] or 0) <= 0:
            return "need_check"
        if not normalize_text(row.get("phone", "")):
            return "no_phone"
        if normalize_text(row.get("last_contacted_at", "")):
            return "contacted"
        return "uncontacted"
    grouped["status_code"] = grouped.apply(calc_company_status, axis=1)
    grouped["status_emoji"] = grouped["status_code"].apply(lambda x: APP_CONFIG["status"][x][0])
    grouped["status_label"] = grouped["status_code"].apply(lambda x: APP_CONFIG["status"][x][1])
    grouped["total_debt_display"] = grouped["total_debt"].apply(format_money)
    grouped["option_label"] = grouped.apply(lambda r: f"{r['status_emoji']} {r['company_name']} | {int(r['ma_tt_count'])} mã | {r['total_debt_display']}", axis=1)
    status_order = {"uncontacted": 1, "no_phone": 2, "contacted": 3, "partial_paid": 4, "need_check": 5, "paid": 6}
    grouped["sort_order"] = grouped["status_code"].map(status_order).fillna(99)
    grouped = grouped.sort_values(["sort_order", "total_debt"], ascending=[True, False]).reset_index(drop=True)
    return grouped

def get_company_interactions(conn, company_key_value):
    return pd.read_sql_query("""
        SELECT result, promised_payment_date, note, created_by, created_at
        FROM interactions
        WHERE company_key = ?
        ORDER BY created_at DESC
    """, conn, params=(company_key_value,))

def get_company_contacts(conn, company_key_value):
    return pd.read_sql_query("""
        SELECT contact_value, contact_person, role, note, created_at
        FROM contacts
        WHERE company_key = ?
        ORDER BY created_at DESC
    """, conn, params=(company_key_value,))

def render_message(period):
    m = APP_CONFIG["message"]
    return m["template"].format(unit_name=m["unit_name"], period=period, website=m["website"], payment_deadline=m["payment_deadline"], staff_name=m["staff_name"], staff_phone=m["staff_phone"])

def get_next_company_key(company_df, current_key):
    candidates = company_df[company_df["status_code"].isin(["uncontacted", "no_phone"])].copy()
    if candidates.empty:
        return current_key
    keys = candidates["company_key"].tolist()
    if current_key in keys:
        idx = keys.index(current_key)
        if idx + 1 < len(keys):
            return keys[idx + 1]
        return keys[0]
    return keys[0]

def export_internal_report_format(lines_df):
    report_headers = [
        "Đưa lưới khóa", "Ngày liên hệ", "Ghi chú", "Mail cập nhật", "SĐT cập nhật",
        "CN TN08", "Tổ Thu cước", "Tuyến kỹ thuật", "Nhân viên thu cước", "SL hđ",
        "Mã Thanh toán", "Tên thanh toán", "Tiền phát sinh", "Số DT liên hệ", "Số seri",
        "Địa chỉ thanh toán", "ghi chú tình trạng", "ghi chú tình trạng", "Tên thanh toán",
        "Tiền phát sinh", "Số DT liên hệ", "Số seri", "Seri", "Địa chỉ thanh toán",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Thuan"
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    yellow_fill = PatternFill("solid", fgColor="FFF200")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, header in enumerate(report_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row_idx, (_, row) in enumerate(lines_df.iterrows(), start=2):
        last_note = normalize_text(row.get("last_note", ""))
        last_result = normalize_text(row.get("last_result", ""))
        promised_date = normalize_text(row.get("promised_payment_date", ""))
        note_parts = []
        if last_result:
            note_parts.append(last_result)
        if promised_date:
            note_parts.append(f"Hẹn: {promised_date}")
        if last_note:
            note_parts.append(last_note)
        final_note = " | ".join(note_parts)
        status_label = ""
        if normalize_text(row.get("last_contacted_at", "")):
            status_label = "Đã liên hệ"
        if float(row.get("paid_amount", 0) or 0) > 0:
            status_label = "Đã đóng"
        excel_values = [
            "", promised_date, final_note, extract_email(final_note), "",
            float(row.get("debt_amount", 0) or 0),
            normalize_text(row.get("collection_team", "")) or APP_CONFIG["defaults"]["collection_team"],
            normalize_text(row.get("route_name", "")),
            normalize_text(row.get("staff_name", "")),
            int(row.get("sl_hd", 1) or 1),
            normalize_text(row.get("ma_tt", "")),
            normalize_text(row.get("company_display", "")),
            float(row.get("generated_amount", 0) or 0),
            normalize_text(row.get("current_phone", "")),
            normalize_text(row.get("serial", "")),
            normalize_text(row.get("address", "")),
            status_label or normalize_text(row.get("old_status_note", "")),
            "", "", "", "", "", "", "",
        ]
        for col_idx, value in enumerate(excel_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_idx == 6:
                cell.fill = yellow_fill
                cell.number_format = '#,##0'
            if col_idx in [13, 20]:
                cell.number_format = '#,##0'
    widths = {"A": 14, "B": 14, "C": 38, "D": 25, "E": 18, "F": 14, "G": 14, "H": 24, "I": 22, "J": 8, "K": 18, "L": 45, "M": 16, "N": 18, "O": 14, "P": 55, "Q": 22, "R": 22, "S": 45, "T": 16, "U": 18, "V": 14, "W": 14, "X": 55}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def export_debt_statement_excel(company_row, company_lines, period):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bang ke cong no"
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.merge_cells("A1:E1")
    ws["A1"] = "THÔNG BÁO CƯỚC VIỄN THÔNG - CNTT"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:E2")
    ws["A2"] = f"THÁNG {period}"
    ws["A2"].font = Font(bold=True, size=14)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A4"] = "Kính gửi:"
    ws["B4"] = company_row.get("company_name", "")
    ws["A4"].font = Font(bold=True)
    ws["B4"].font = Font(bold=True)
    ws["E5"] = "ĐVT: VNĐ"
    ws["E5"].font = Font(italic=True)
    ws["E5"].alignment = Alignment(horizontal="right")
    headers = ["TT", "Mã Thanh toán", "Tên KH", "Địa chỉ KH", "Tiền cước dịch vụ"]
    start_row = 6
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    total = 0
    for idx, (_, row) in enumerate(company_lines.iterrows(), start=1):
        excel_row = start_row + idx
        tien = float(row.get("debt_amount", 0) or 0)
        total += tien
        values = [idx, row.get("ma_tt", ""), row.get("company_display", ""), row.get("address", ""), tien]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_idx == 5:
                cell.number_format = '#,##0'
    total_row = start_row + len(company_lines) + 1
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    ws.cell(row=total_row, column=1, value="Tổng")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=5, value=total)
    ws.cell(row=total_row, column=5).font = Font(bold=True)
    ws.cell(row=total_row, column=5).number_format = '#,##0'
    for col in range(1, 6):
        ws.cell(row=total_row, column=col).border = border
    note_row = total_row + 2
    ws.cell(row=note_row, column=1, value="Quý khách hàng vui lòng thanh toán trước ngày 15 hàng tháng.")
    ws.cell(row=note_row, column=1).font = Font(color="FF0000")
    ws.cell(row=note_row + 1, column=1, value="Quá hạn thanh toán cước thiết bị thông tin của Quý khách sẽ tạm ngưng hoạt động.")
    ws.cell(row=note_row + 1, column=1).font = Font(color="FF0000")
    ws.cell(row=note_row + 3, column=1, value="Quý khách hàng vui lòng thanh toán qua tài khoản sau:")
    ws.cell(row=note_row + 4, column=1, value="- Tên tài khoản: VIỄN THÔNG ĐỒNG NAI - TẬP ĐOÀN BƯU CHÍNH VIỄN THÔNG VIỆT NAM")
    ws.cell(row=note_row + 5, column=1, value="- Số tài khoản: 0121000890707 tại Ngân hàng Vietcombank - Chi nhánh Đồng Nai")
    ws.cell(row=note_row + 6, column=1, value=f"- Nội dung: {company_row.get('company_name', '')} thanh toán cước trả sau")
    ws.cell(row=note_row + 6, column=1).font = Font(color="FF0000")
    widths = {"A": 8, "B": 22, "C": 42, "D": 55, "E": 20}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

st.set_page_config(page_title="Mini-CRM Thu Cước KHDN", page_icon="📞", layout="wide")
conn = get_conn()
init_db(conn)
st.title(APP_CONFIG["app_title"])
st.caption("Data theo MA_TT, thao tác theo Công ty. Lịch sử gọi lưu theo company_key.")

tabs = st.tabs(["1. Nạp dữ liệu", "2. Bàn làm việc theo công ty", "3. Báo cáo", "4. Quản trị dữ liệu"])

with tabs[0]:
    st.subheader("1. Nạp dữ liệu")
    st.info("Quy trình: DS giao kỳ cước → lọc nhân viên → TN08 → lấy Total Nợ thu vét → gom theo công ty.")
    assign_file = st.file_uploader("A. Upload DS giao kỳ cước", type=["xlsx", "xls", "csv"], key="assign")
    df_assignment = None
    selected_staff = None
    if assign_file:
        try:
            if assign_file.name.lower().endswith((".xlsx", ".xls")):
                sheets = list_sheets(assign_file)
                default_idx = 0
                for i, s in enumerate(sheets):
                    if normalize_key(s) in ["ds", "thuan", "giao"]:
                        default_idx = i
                        break
                sheet = st.selectbox("Chọn sheet DS giao", sheets, index=default_idx)
                raw_assign = read_file(assign_file, sheet)
            else:
                raw_assign = read_file(assign_file)
            df_assignment = standardize_assignment(raw_assign, fallback_staff=APP_CONFIG["default_staff"])
            staff_list = sorted(df_assignment["staff_name"].dropna().unique().tolist())
            default_idx = 0
            for i, s in enumerate(staff_list):
                if staff_key(s) == staff_key(APP_CONFIG["default_staff"]):
                    default_idx = i
                    break
            selected_staff = st.selectbox("Chọn nhân viên cần lọc", staff_list, index=default_idx)
            route_df = df_assignment[df_assignment["staff_key"] == staff_key(selected_staff)].copy()
            st.success(f"Đã lọc nhân viên {selected_staff}: {len(route_df)} mã thanh toán.")
            st.dataframe(route_df[["staff_name", "ma_tt", "customer_name", "phone", "address", "generated_amount"]].head(100), use_container_width=True)
        except Exception as e:
            st.error(f"Lỗi đọc DS giao kỳ cước: {e}")
    st.divider()
    tn08_file = st.file_uploader("B. Upload TN08 hóa đơn chưa thu", type=["xlsx", "xls", "csv"], key="tn08")
    period = st.text_input("Kỳ cước cho mẫu tin/bảng kê", value=APP_CONFIG["message"]["default_period"])
    if tn08_file and df_assignment is not None and selected_staff:
        try:
            if tn08_file.name.lower().endswith((".xlsx", ".xls")):
                sheets = list_sheets(tn08_file)
                default_idx = 0
                for i, s in enumerate(sheets):
                    if "tn08" in normalize_key(s):
                        default_idx = i
                        break
                tn08_sheet = st.selectbox("Chọn sheet TN08", sheets, index=default_idx)
                raw_tn08 = read_file(tn08_file, tn08_sheet)
            else:
                raw_tn08 = read_file(tn08_file)
            df_tn08 = standardize_tn08(raw_tn08)
            working = build_working_dataset(df_assignment, df_tn08, selected_staff)
            if working.empty:
                st.warning("Không có mã nào thuộc nhân viên đã chọn.")
            else:
                company_count = working["company_key"].nunique()
                st.success(f"Đã dò TN08 xong. Tạo được {company_count} công ty/nhóm từ {len(working)} mã.")
                st.metric("Tổng tiền cần thu theo TN08", format_money(working["debt_amount"].sum()))
                st.dataframe(working[["company_name", "ma_tt", "phone", "debt_amount_display", "generated_amount_display", "data_status"]].head(200), use_container_width=True)
                if st.button("Lưu tuyến này vào CRM"):
                    batch_id = create_batch(conn, "working_dataset", f"{assign_file.name} + {tn08_file.name}", selected_staff, len(working), float(working["debt_amount"].sum()))
                    insert_work_items(conn, batch_id, working)
                    st.success(f"Đã lưu vào CRM. Batch ID: {batch_id}. Sang tab 2 để gọi theo công ty.")
        except Exception as e:
            st.error(f"Lỗi xử lý TN08: {e}")
    st.divider()
    paid_file = st.file_uploader("C. Upload file khách đã đóng nếu có", type=["xlsx", "xls", "csv"], key="paid")
    if paid_file:
        try:
            raw_paid = read_file(paid_file)
            paid = standardize_paid(raw_paid)
            st.dataframe(paid.head(100), use_container_width=True)
            if st.button("Lưu danh sách đã đóng"):
                batch_id = create_batch(conn, "paid", paid_file.name, "", len(paid), float(paid["paid_amount"].sum()))
                insert_paid_updates(conn, batch_id, paid)
                st.success(f"Đã lưu danh sách đã đóng. Batch ID: {batch_id}")
        except Exception as e:
            st.error(f"Lỗi đọc file đã đóng: {e}")

with tabs[1]:
    st.subheader("2. Bàn làm việc theo công ty")
    lines_df = get_current_lines(conn)
    company_df = build_company_view(lines_df)
    if company_df.empty:
        st.info("Chưa có dữ liệu. Hãy nạp DS giao + TN08 ở tab 1.")
    else:
        if "selected_company_key" not in st.session_state:
            st.session_state.selected_company_key = company_df.iloc[0]["company_key"]
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Số công ty", len(company_df))
        c2.metric("Tổng tiền còn thu", format_money(company_df[~company_df["status_code"].isin(["paid"])]["total_debt"].sum()))
        c3.metric("Chưa liên hệ", int((company_df["status_code"] == "uncontacted").sum()))
        c4.metric("Đã liên hệ", int((company_df["status_code"] == "contacted").sum()))
        c5.metric("Thiếu số", int((company_df["status_code"] == "no_phone").sum()))
        status_options = ["Tất cả"] + sorted(company_df["status_label"].unique().tolist())
        status_filter = st.selectbox("Lọc trạng thái công ty", status_options)
        q = st.text_input("Tìm tên công ty / số điện thoại")
        view = company_df.copy()
        if status_filter != "Tất cả":
            view = view[view["status_label"] == status_filter]
        if q:
            qk = remove_accents(q)
            view = view[view["company_name"].apply(remove_accents).str.contains(qk, na=False) | view["phone"].apply(remove_accents).str.contains(qk, na=False)]
        show_cols = ["status_emoji", "status_label", "company_name", "ma_tt_count", "total_debt_display", "phone", "last_result", "promised_payment_date", "last_contacted_at"]
        st.dataframe(view[show_cols], use_container_width=True, height=320)
        if view.empty:
            st.warning("Không có công ty phù hợp bộ lọc.")
        else:
            option_labels = view["option_label"].tolist()
            keys = view["company_key"].tolist()
            if st.session_state.selected_company_key in keys:
                default_idx = keys.index(st.session_state.selected_company_key)
            else:
                default_idx = 0
                st.session_state.selected_company_key = keys[0]
            selected_label = st.selectbox("Chọn công ty để xử lý", option_labels, index=default_idx)
            selected_company_key = keys[option_labels.index(selected_label)]
            st.session_state.selected_company_key = selected_company_key
            company_row = company_df[company_df["company_key"] == selected_company_key].iloc[0].to_dict()
            company_lines = lines_df[lines_df["company_key"] == selected_company_key].copy()
            st.divider()
            left, right = st.columns([1.15, 1])
            with left:
                st.markdown("### Chi tiết công ty")
                st.write(f"**Công ty:** {company_row.get('company_name', '')}")
                st.write(f"**Số điện thoại/Zalo ưu tiên:** {company_row.get('phone', '') or 'Chưa có'}")
                st.write(f"**Số mã thanh toán:** {int(company_row.get('ma_tt_count') or 0)}")
                st.write(f"**Tổng tiền cần thu:** {company_row.get('total_debt_display', '')}")
                st.write(f"**Trạng thái:** {company_row.get('status_emoji', '')} {company_row.get('status_label', '')}")
                st.markdown("#### Danh sách mã thanh toán của công ty")
                detail_cols = ["ma_tt", "company_display", "current_phone", "address", "debt_amount_display", "generated_amount_display", "data_status"]
                st.dataframe(company_lines[detail_cols], use_container_width=True, height=220)
                st.markdown("#### Lịch sử gọi công ty")
                st.dataframe(get_company_interactions(conn, selected_company_key), use_container_width=True, height=160)
                st.markdown("#### Lịch sử số liên hệ")
                st.dataframe(get_company_contacts(conn, selected_company_key), use_container_width=True, height=140)
            with right:
                st.markdown("### Cập nhật kết quả cuộc gọi")
                after_save_mode = st.radio("Sau khi lưu", ["Giữ nguyên công ty hiện tại", "Tự chuyển sang công ty tiếp theo chưa liên hệ"], index=1)
                with st.form("call_form", clear_on_submit=True):
                    result = st.selectbox("Kết quả cuộc gọi", APP_CONFIG["contact_results"])
                    promised = st.text_input("Ngày hẹn thanh toán", placeholder="VD: 15/06")
                    note = st.text_area("Ghi chú")
                    created_by = st.text_input("Người cập nhật", value=APP_CONFIG["message"]["staff_name"])
                    submit = st.form_submit_button("Lưu lịch sử gọi")
                    if submit:
                        add_interaction(conn, selected_company_key, result, promised, note, created_by)
                        if after_save_mode.startswith("Tự chuyển"):
                            next_key = get_next_company_key(company_df, selected_company_key)
                            st.session_state.selected_company_key = next_key
                        else:
                            st.session_state.selected_company_key = selected_company_key
                        st.success("Đã lưu lịch sử gọi.")
                        st.rerun()
                st.markdown("### Cập nhật số Zalo/SĐT mới")
                with st.form("contact_form", clear_on_submit=True):
                    new_phone = st.text_input("Số Zalo/SĐT mới")
                    person = st.text_input("Tên người phụ trách")
                    role = st.text_input("Vai trò", value="Kế toán")
                    contact_note = st.text_area("Ghi chú số liên hệ")
                    submit_contact = st.form_submit_button("Lưu số liên hệ")
                    if submit_contact:
                        phone = normalize_phone(new_phone)
                        if not phone:
                            st.error("Số điện thoại chưa hợp lệ.")
                        else:
                            add_contact(conn, selected_company_key, phone, person, role, contact_note)
                            st.success("Đã lưu số liên hệ mới.")
                            st.rerun()
            st.divider()
            st.markdown("### Mẫu tin nhắn và bảng kê công nợ")
            period_msg = st.text_input("Kỳ cước trong tin nhắn/bảng kê", value=APP_CONFIG["message"]["default_period"], key="period_msg")
            msg = render_message(period_msg)
            st.code(msg, language="text")
            col_msg, col_statement = st.columns(2)
            with col_msg:
                if st.button("Ghi nhận đã copy/gửi tin"):
                    add_sent_message(conn, selected_company_key, msg, APP_CONFIG["message"]["staff_name"])
                    st.success("Đã ghi nhận.")
            with col_statement:
                statement_bytes = export_debt_statement_excel(company_row, company_lines, period_msg)
                file_safe = normalize_key(company_row.get("company_name", "bang_ke"))
                st.download_button("Tải bảng kê công nợ công ty", data=statement_bytes, file_name=f"bang_ke_cong_no_{file_safe}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

with tabs[2]:
    st.subheader("3. Báo cáo")
    lines_df = get_current_lines(conn)
    company_df = build_company_view(lines_df)
    if lines_df.empty:
        st.info("Chưa có dữ liệu.")
    else:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Số công ty", len(company_df))
        c2.metric("Tổng mã", len(lines_df))
        c3.metric("Tổng tiền còn thu", format_money(company_df[~company_df["status_code"].isin(["paid"])]["total_debt"].sum()))
        c4.metric("Đã liên hệ", int((company_df["status_code"] == "contacted").sum()))
        st.markdown("### Báo cáo theo công ty")
        st.dataframe(company_df, use_container_width=True, height=320)
        st.markdown("### Báo cáo chi tiết theo Mã thanh toán")
        st.dataframe(lines_df, use_container_width=True, height=360)
        st.download_button("Tải báo cáo nội bộ đúng mẫu Excel", data=export_internal_report_format(lines_df), file_name="bao_cao_thu_cuoc_dung_mau.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

with tabs[3]:
    st.subheader("4. Quản trị dữ liệu")
    st.markdown("### Lịch sử import")
    batches = pd.read_sql_query("""
        SELECT batch_id, import_type, file_name, staff_filter, row_count, total_amount, created_at
        FROM import_batches
        ORDER BY batch_id DESC
        LIMIT 50
    """, conn)
    st.dataframe(batches, use_container_width=True)
    st.markdown("### Reset database")
    st.warning("Bản v2 dùng database mới. Nếu muốn test lại từ đầu, gõ RESET rồi xóa dữ liệu.")
    confirm = st.text_input("Gõ RESET để xác nhận")
    if st.button("Xóa toàn bộ dữ liệu") and confirm == "RESET":
        conn.executescript("""
            DELETE FROM sent_messages;
            DELETE FROM contacts;
            DELETE FROM interactions;
            DELETE FROM paid_updates;
            DELETE FROM work_items;
            DELETE FROM import_batches;
        """)
        conn.commit()
        st.success("Đã reset dữ liệu.")
        st.rerun()
